# -*- coding: utf-8 -*-
"""数据加载与清洗"""
import numpy as np
import pandas as pd
from io import BytesIO


def _map_columns(df):
    """列名映射：模糊匹配源列名到标准列名"""
    col_map = {}
    for c in df.columns:
        cl = str(c).lower().strip()
        if any(k in cl for k in ["日期", "send", "date"]):
            col_map[c] = "发送日期"
        elif any(k in cl for k in ["计划类型", "plan_type"]):
            col_map[c] = "计划类型"
        elif any(k in cl for k in ["渠道", "channel"]):
            col_map[c] = "渠道"
        elif "plan_id" in cl or "plan id" in cl:
            col_map[c] = "plan_id"
        elif any(k in cl for k in ["plan名称", "plan_name", "plan 名称"]):
            col_map[c] = "plan名称"
        elif "owner" in cl:
            col_map[c] = "owner"
        elif any(k in cl for k in ["是否用券", "coupon"]):
            col_map[c] = "是否用券"
        elif any(k in cl for k in ["预计触达", "exp_reach"]):
            col_map[c] = "预计触达"
        elif any(k in cl for k in ["触达成功", "reach"]):
            col_map[c] = "触达成功"
        elif any(k in cl for k in ["点击人次", "click"]) and "下单" not in cl:
            col_map[c] = "点击人次"
        elif any(k in cl for k in ["点击后下单", "post_click"]):
            col_map[c] = "点击后下单人次"
        elif cl in ["gc", "订单gc", "order_gc", "订单 gc"] or cl == "gc":
            col_map[c] = "订单GC"
        elif cl in ["sales", "订单sales", "order_sales", "订单 sales"] or cl == "sales":
            col_map[c] = "订单Sales"
    df.rename(columns=col_map, inplace=True)
    return df


def compute_ctr(df, click_col="点击人次", reach_col="触达成功"):
    """CTR = 点击 / 触达 * 100，除零返回 0，保留 2 位小数"""
    reach = df[reach_col].astype(float).replace(0, np.nan)
    return (df[click_col] / reach * 100).round(2).fillna(0)


def _convert_types(df):
    """类型转换 + 派生指标 + 必需列校验"""
    if "发送日期" in df.columns:
        df["发送日期"] = pd.to_datetime(df["发送日期"], errors="coerce")
    for col in ["触达成功", "点击人次", "点击后下单人次", "订单GC", "订单Sales", "预计触达"]:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0).astype(int)
    if "触达成功" in df.columns and "点击人次" in df.columns:
        df["CTR"] = compute_ctr(df)
    else:
        df["CTR"] = 0
    required = ["发送日期", "渠道", "计划类型", "触达成功", "点击人次", "plan_id", "owner"]
    missing = [c for c in required if c not in df.columns]
    if missing:
        raise ValueError(f"缺少必需列（映射后）: {', '.join(missing)}。请检查文件格式。")
    return df


def read_dau_sheet(file_bytes):
    """读取 XLSX 第二个 sheet（按天去重 DAU，可选分渠道）。

    支持两种格式（自动检测）：
      - 新格式（≥3 列且第 2 列非空）：日期 / 渠道 / DAU；"ALL"/"all" 表示总 DAU
      - 旧格式（2 列）：日期 / DAU；解析后渠道列填充为 "ALL"

    返回列：日期、渠道、DAU。sheet 不足 2 个或数据为空时返回空 DataFrame。
    接受 bytes（调用方从 uploaded.read() 获得）。
    """
    import openpyxl
    wb = openpyxl.load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
    try:
        if len(wb.sheetnames) < 2:
            return pd.DataFrame()
        ws = wb[wb.sheetnames[1]]
        rows = list(ws.iter_rows(values_only=True))
    finally:
        wb.close()

    if len(rows) < 2:
        return pd.DataFrame()

    # 解析：3 列含非空渠道 → 新格式；否则 2 列 → 旧格式（渠道置 ALL）
    parsed = []
    for r in rows[1:]:
        if len(r) >= 3 and r[0] is not None and r[1] is not None and r[2] is not None:
            parsed.append((r[0], str(r[1]).strip(), r[2]))
        elif len(r) >= 2 and r[0] is not None and r[1] is not None:
            parsed.append((r[0], "ALL", r[1]))
    if not parsed:
        return pd.DataFrame()

    df = pd.DataFrame(parsed, columns=["日期_raw", "渠道", "DAU_raw"])
    df["日期"] = pd.to_datetime(df["日期_raw"], errors="coerce")
    df["DAU"] = pd.to_numeric(df["DAU_raw"], errors="coerce")
    df = df.dropna(subset=["日期", "DAU"])[["日期", "渠道", "DAU"]]
    return df.reset_index(drop=True)


def load_csv(uploaded_file):
    """读取 CSV，仅保留 A-M 列（0-12），跳过 JSON 列"""
    data = uploaded_file.read()
    encodings = ["utf-8", "utf-8-sig", "gbk", "gb2312", "latin1"]
    df = None
    for enc in encodings:
        try:
            df = pd.read_csv(BytesIO(data), encoding=enc, on_bad_lines="skip")
            break
        except Exception:
            continue
    if df is None:
        raise ValueError("无法解析 CSV 文件")
    # 只保留前 13 列 (A-M)，跳过 JSON 列
    if df.shape[1] >= 13:
        df = df.iloc[:, :13]
    df = _map_columns(df)
    df = _convert_types(df)
    return df


def load_xlsx(uploaded_file):
    """读取 XLSX，仅保留 A-M 列（0-12），跳过 JSON 列，完整保留 emoji"""
    import openpyxl
    wb = openpyxl.load_workbook(BytesIO(uploaded_file.read()), read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if len(rows) < 2:
        raise ValueError("XLSX 文件没有数据行")

    headers = [str(h).strip() if h else '' for h in rows[0]]
    data_rows = rows[1:]

    df = pd.DataFrame(data_rows, columns=headers)
    # 只保留前 13 列 (A-M)，跳过 JSON 列
    if df.shape[1] >= 13:
        df = df.iloc[:, :13]
    df = _map_columns(df)
    df = _convert_types(df)
    return df
